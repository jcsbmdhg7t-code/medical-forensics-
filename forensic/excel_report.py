#!/usr/bin/env python3
"""
Excel rapport generator — Grothe forensisch dossier
Produceert een dagelijks bijgewerkt .xlsx met alle bevindingen.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import datetime

ROOT = Path(__file__).parent.parent
REPORTS = ROOT / "reports"

# Kleurpalet
RED    = "C0392B"; ORANGE = "E67E22"; YELLOW = "F1C40F"
GREEN  = "27AE60"; BLUE   = "2980B9"; GREY   = "7F8C8D"
WHITE  = "FFFFFF"; DARK   = "2C3E50"; LIGHT  = "ECF0F1"

BEWIJS_COLOR = {"H": RED, "HOOG": RED, "MIDDEL": ORANGE, "LAAG": YELLOW, "PENDING": GREY}

def hdr(ws, row, col, text, bold=True, bg=DARK, fg=WHITE, size=10):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(bold=bold, color=fg, size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    return cell

def val(ws, row, col, text, bg=None, bold=False, wrap=True):
    cell = ws.cell(row=row, column=col, value=str(text) if text else "")
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=bold, size=9)
    cell.alignment = Alignment(wrap_text=wrap, vertical="top")
    return cell

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def freeze(ws, cell="A2"):
    ws.freeze_panes = cell

# ── Sheet builders ─────────────────────────────────────────────────────────

def sheet_dashboard(wb, index, today):
    ws = wb.create_sheet("📊 Dashboard", 0)
    ws.sheet_view.showGridLines = False
    hdr(ws, 1, 1, f"FORENSISCH DOSSIER GROTHE — Dagrapport {today}", size=14, bg="1A252F")
    ws.merge_cells("A1:F1")
    ws.row_dimensions[1].height = 28

    stats = [
        ("Geanalyseerde bestanden", len(index.get("entries", {}))),
        ("Unieke actoren", len(set(a for e in index["entries"].values() for a in e.get("actors",[])))),
        ("Open todo's", len(index.get("todos", []))),
        ("NB-nummers gevonden", len(set(nb for e in index["entries"].values() for nb in e.get("nb_refs",[])))),
        ("Causale verbanden", len(index.get("causal_map", []))),
        ("Hypotheses", len(index.get("hypotheses", []))),
    ]
    hdr(ws, 3, 1, "STATISTIEKEN", bg=DARK); ws.merge_cells("A3:B3")
    for i, (label, count) in enumerate(stats, start=4):
        val(ws, i, 1, label, bg=LIGHT)
        val(ws, i, 2, count, bold=True)

    hdr(ws, 3, 4, "KRITIEKE DEADLINES", bg=RED); ws.merge_cells("D3:F3")
    deadlines = [
        ("28-06-2026", "KRITIEK", "Parnassia portaaltoegang + logging vervalt"),
        ("22-07-2026", "HOOG",    "Zitting Rb Noord-Holland C/15/376914"),
        ("03-08-2026", "HOOG",    "Hof van Discipline 260153 uitspraak"),
    ]
    for i, (dl, prio, desc) in enumerate(deadlines, start=4):
        bg = RED if prio == "KRITIEK" else ORANGE
        val(ws, i, 4, dl, bg=bg, bold=True)
        val(ws, i, 5, prio, bg=bg)
        val(ws, i, 6, desc)

    set_col_widths(ws, [25, 12, 5, 12, 10, 45])

def sheet_index(wb, index):
    ws = wb.create_sheet("📁 Index")
    headers = ["Bestand", "Type", "Grootte", "Bewijs", "NB-refs", "Spoor", "Actoren", "Juridische gronden", "Geanalyseerd", "Hash"]
    for c, h in enumerate(headers, 1):
        hdr(ws, 1, c, h)
    freeze(ws)

    entries = sorted(index.get("entries", {}).values(),
                     key=lambda x: BEWIJS_COLOR.get(x.get("bewijs_weight","PENDING"), GREY),
                     reverse=True)

    for r, e in enumerate(entries, start=2):
        bg_row = LIGHT if r % 2 == 0 else WHITE
        bw = e.get("bewijs_weight", "PENDING")
        bw_bg = BEWIJS_COLOR.get(bw, GREY)

        val(ws, r, 1, Path(e["file"]).name, bg=bg_row)
        val(ws, r, 2, Path(e["file"]).suffix.upper(), bg=bg_row)
        val(ws, r, 3, f"{e.get('size_kb',0)} KB", bg=bg_row)
        val(ws, r, 4, bw, bg=bw_bg, bold=True)
        val(ws, r, 5, ", ".join(e.get("nb_refs", [])[:8]), bg=bg_row)
        val(ws, r, 6, e.get("spoor", ""), bg=bg_row)
        val(ws, r, 7, "; ".join(e.get("actors", [])[:3]), bg=bg_row)
        val(ws, r, 8, "; ".join(e.get("legal_refs", [])[:4]), bg=bg_row)
        val(ws, r, 9, e.get("analyzed", ""), bg=bg_row)
        val(ws, r, 10, e.get("hash", ""), bg=bg_row)

    set_col_widths(ws, [35, 6, 8, 8, 30, 20, 40, 50, 12, 10])

def sheet_actoren(wb, index):
    ws = wb.create_sheet("👤 Actoren")
    headers = ["Naam", "AGB/ID", "Rol", "Vektis-discrepantie", "Betrokken in NB's", "Spoor", "Cui-bono status"]
    for c, h in enumerate(headers, 1):
        hdr(ws, 1, c, h)
    freeze(ws)

    ACTOREN = [
        ("A. al-Mousawi",         "84126524",       "Basisarts va. 04-09-2023",
         "Entry 13-01-2020 (3,5jr te vroeg); SUBSTANCEHXQNR zonder BIG",
         "NB-108, NB-169", "Spaarne", "Tweemansprocedure 02-10-2024 — kan zelf slachtoffer zijn"),
        ("J.P.J. van der List",   "84115003",       "Basisarts va. 05-06-2020; geen ortho",
         "AIOS orthopedie label 29-11-2018 (1,5jr te vroeg)",
         "NB-148", "Spaarne", "Pending Vektis — cui-bono: dekmantel AIOS-label"),
        ("N.M. Nota",             "84107660/51504662","Uitsluitend basisarts, nooit reum.",
         "AIOS reumatologie label; massa-sluiting 5 diagnoses 02-10-2024; geen supervisie",
         "NB-108, NB-171", "Spaarne", "Tweemansprocedure — kan zelf slachtoffer zijn"),
        ("J.M. Blauw",            "PM (Vektis pend.)","Basisarts (datum pending)",
         "Laborders 29-11-2019 zonder gedocumenteerd consult",
         "NB-173", "Spaarne", "Pending Vektis — dekmantel voor laborders?"),
        ("M. van der Kroon (MRK)","FACILITAIR",      "Facilitair admin",
         "Toegangslog 14-12-2017 + override.css developer 2024-2025",
         "NB-168, CC-J", "Spaarne", "Directe technische betrokkenheid bewezen"),
        ("K.J. Burlage",          "67480",           "Huisarts (deceased 15-02-2024)",
         "Naam post-mortem ingevoerd als GP op 03-12-2017",
         "NB-36, NB-170", "Huisarts", "Deceased — naam misbruikt na overlijden"),
        ("D. Kox",                "KNO-arts",        "KNO, Vijf Meren",
         "Bronbrief 04-12-2017: diagnose = OTRIVINISME (niet F19.1)",
         "NB-116", "Spaarne", "Getuige/slachtoffer: zijn brief werd hergecodeerd"),
        ("Ext=999999",            "ANONIEM",         "Anonieme auteur EPD",
         "6× in DOC0010; nachtelijke batch 04:34 CET op AVG-bevriezingsdatum",
         "NB-166, NB-18", "Beide", "Onbekend — batch-operatie geautomatiseerd"),
        ("Hoppinger B.V.",        "Leverancier",     "Webontwikkelaar portaal",
         "spaarne-rebuild.productie.hoppinger.com in productie hardcoded",
         "NB-114", "Spaarne", "Supply-chain risico; eigen belang: contract behoud"),
    ]

    for r, actor in enumerate(ACTOREN, start=2):
        bg = LIGHT if r % 2 == 0 else WHITE
        for c, v in enumerate(actor, start=1):
            val(ws, r, c, v, bg=bg)

    set_col_widths(ws, [25, 18, 28, 50, 20, 12, 50])

def sheet_tijdlijn(wb, index):
    ws = wb.create_sheet("📅 Tijdlijn")
    headers = ["Datum", "Tijd", "Event", "Actor", "Systeem", "NB-ref", "Juridische betekenis", "Bewijs"]
    for c, h in enumerate(headers, 1):
        hdr(ws, 1, c, h)
    freeze(ws)

    TIJDLIJN = [
        ("24-03-2016","","Xylometazoline neusspray voorgeschreven (pre-existent)","Van der Meij","Medicom","NB-131","Rhinitis pre-existent vóór F19-registratie","HOOG"),
        ("04-12-2017","09:38","KNO-brief D. Kox: diagnose OTRIVINISME","D. Kox","Vijf Meren","NB-116","F19.1 is hercodering van KNO-diagnose","HOOG"),
        ("04-12-2017","~","F19.1 geregistreerd in Spaarne Epic","Ext=999999","Epic EPD","NB-01","Zelfde dag als KNO-brief; Isabel mogelijk niet aanwezig","HOOG"),
        ("05-12-2017","11:00","Psychiater Tusenius: GEEN ADHD/drugs","A. Tusenius","PsyQ","NB-124","Tegenbewijs: dag ná F19-registratie","HOOG"),
        ("06-12-2017","","Burlage verwijsbrief GGZ: GEEN F19","K.J. Burlage","Medicom","NB-128","Intern tegenbewijs uit dezelfde praktijk","HOOG"),
        ("14-12-2017","","MRK (Facilitair Admin) toegangslog","M. van der Kroon","Epic SG","NB-168","CC-J: zelfde persoon als override.css developer 2024","HOOG"),
        ("28-02-2018","","PsyQ DSM: uitsluitend 314.01 ADHD, GEEN F19","C. Gulikers/P. Bisschop","PsyQ","NB-130","PsyQ behandelde 2018-2019 zonder F19","HOOG"),
        ("23-11-2018","","'GEEN dexamfetamine meer leveren!' (Van der Meij)","M.I. van der Meij","Medicom","NB-129","Medicatie via Tusenius/Burlage, niet Spaarne","HOOG"),
        ("29-11-2018","","Van der List: AIOS orthopedie label in EPD","J.P.J. van der List","Epic SG","NB-148","Basisarts pas 05-06-2020 — 1,5 jaar te vroeg","HOOG"),
        ("13-01-2020","","Al-Mousawi: co-ass. toegangslog / SEH-entry EPD","A. al-Mousawi","Epic SG","NB-169","Basisarts pas 04-09-2023 — 3,5 jaar te vroeg","HOOG"),
        ("29-11-2019","","J.M. Blauw: 4 laborders zonder consult","J.M. Blauw","Epic SG","NB-173","Geen consult in bezoekenoverzicht; basisarts pending","MIDDEL"),
        ("02-09-2024","","Nota-relatie Spaarne start","N.M. Nota","Epic SG","NB-171","1 maand vóór massa-sluiting","HOOG"),
        ("02-10-2024","07:54","Nota sluit 5 diagnoses in 31 sec","N.M. Nota","Epic SG","NB-04/94","Tweemansprocedure — geen reum. supervisie","HOOG"),
        ("02-10-2024","07:54","Al-Mousawi: SNOMED 361055000 via SUBSTANCEHXQNR","A. al-Mousawi","Epic SG","NB-108","Zelfde dag; functioneel equivalent F19 via andere module","HOOG"),
        ("10-01-2026","03:34","Nachtelijke anon. auteur in 7/12 CDA-documenten","Ext=999999","Epic SG","NB-166","Dag van AVG-art.18-verzoek; batchproces","HOOG"),
        ("11-01-2026","23:55","Medicom batch sweep alle 7 dossier-categorieën","?","Medicom","NB-157","Dag ná AVG-bevriezingsdatum","HOOG"),
        ("20-01-2026","","BI EPIC GECERTIFICEERD actor in bevroren dossier","J. de G.","Epic SG","NB-100/156","9 dagen ná AVG-art.18-verzoek","HOOG"),
    ]

    for r, row in enumerate(TIJDLIJN, start=2):
        bg = LIGHT if r % 2 == 0 else WHITE
        bw_color = BEWIJS_COLOR.get(row[7], GREY)
        for c, v in enumerate(row, start=1):
            cell_bg = bw_color if c == 8 else bg
            val(ws, r, c, v, bg=cell_bg)

    set_col_widths(ws, [12, 8, 55, 22, 14, 16, 50, 8])

def sheet_causaal(wb, index):
    ws = wb.create_sheet("🔗 Causaal")
    headers = ["Van (event/datum)", "Naar (event/datum)", "Type verband", "Bewijs", "Juridische basis", "NB-refs", "Verificatie status"]
    for c, h in enumerate(headers, 1):
        hdr(ws, 1, c, h)
    freeze(ws)

    CAUSAAL = [
        ("F19.1 registratie 04-12-2017", "IBS-poging december 2018",
         "Directe causaliteit: stigma → coercieve maatregel",
         "KNO-brief (NB-116) + PsyQ tegenbewijs (NB-131) bewijzen F19 onjuist was",
         "Art. 6:162 BW; Art. 5 ECHR; Art. 282 Sr", "NB-116/131/133", "BEVESTIGD — CC-A"),
        ("Al-Mousawi entry 13-01-2020", "Nota massa-sluiting 02-10-2024",
         "Mogelijke coördinatie: SEH-toegang → diagnose-manipulatie",
         "Beide op zelfde ext-ID; SUBSTANCEHXQNR toegang zonder behandelrel.",
         "Art. 225 Sr; BIG art. 36", "NB-108/160/169", "HYPOTHESIS — verificatie pending"),
        ("F19.1 (Spaarne)", "228 data-requests Brijder/Indigo",
         "SNOMED 228273003 seeding: vervalst label gesynchroniseerd via Parnassia-netwerk",
         "MedMij-portabiliteitsverslag (NB-113)",
         "Art. 138ab Sr; Art. 272 Sr; AVG art. 5", "NB-113", "BEVESTIGD — externe sync"),
        ("MRK Facilitair Admin 14-12-2017", "override.css hiding 2024-2025",
         "Zelfde persoon 7 jaar later technische verberging patiëntdata",
         "Toegangslog + ingebrekestelling v22 (NB-168)",
         "Art. 225 Sr; NEN 7510", "NB-168/CC-J", "BEVESTIGD — CC-J"),
        ("Burlage naam 03-12-2017", "Burlage deceased 15-02-2024",
         "Post-mortem naam-invoer: naam gebruikt vóór en ná overlijden",
         "Overlijdensdatum vs. dossier-invoerdatum (NB-36)",
         "Art. 225 Sr; WGBO 7:454", "NB-36/170", "BEVESTIGD"),
        ("AVG art.18 verzoek 10-01-2026 12:00", "Nachtelijke batch 10-01-2026 03:34 UTC",
         "Reactieve manipulatie: wijziging ná ontvangst bevriezingsverzoek",
         "Timestamp verzoek vs. CDA-auteur timestamp (NB-166)",
         "AVG art. 18; AVG art. 5(1)(f); Art. 225 Sr", "NB-166", "BEVESTIGD"),
        ("spaarne-rebuild.productie.hoppinger.com", "Live Epic productiedatabase",
         "Supply chain: testomgeving gekoppeld aan live patiëntendata",
         "Hardcoded URL in portaalconfig (NB-114)",
         "AVG art. 28/32; NEN 7510 §12.1.4", "NB-114", "BEVESTIGD"),
    ]

    for r, row in enumerate(CAUSAAL, start=2):
        bg = LIGHT if r % 2 == 0 else WHITE
        status = row[6]
        status_bg = GREEN if "BEVESTIGD" in status else (ORANGE if "HYPOTHESIS" in status else GREY)
        for c, v in enumerate(row, start=1):
            cell_bg = status_bg if c == 7 else bg
            val(ws, r, c, v, bg=cell_bg)

    set_col_widths(ws, [30, 30, 35, 50, 35, 20, 25])

def sheet_hypotheses(wb, index):
    ws = wb.create_sheet("💡 Hypotheses")
    headers = ["Hypothese", "Bewijs VOOR", "Bewijs TEGEN", "Verificatie actie", "Prioriteit", "Verwacht scenario"]
    for c, h in enumerate(headers, 1):
        hdr(ws, 1, c, h)
    freeze(ws)

    HYPOTHESES = [
        ("F19.1 is doelbewust geplaatst als juridisch wapen om GGZ-dwangmaatregelen te legitimeren",
         "KNO-brief bewijst otrivinisme; PsyQ bevestigt nooit F19; Burlage brief geen F19; datum-coincidentie",
         "Mogelijk administratieve fout KNO → EPD codering (minder waarschijnlijk gezien driepatroon)",
         "Originele papieren KNO-dossier opvragen bij Vijf Meren",
         "HOOG", "Gefabriceerde indicatie → IBS-poging → stigmatisering → behandelbelemmering"),
        ("De vier AIOS/basisartsen (Mousawi, Van der List, Nota, Blauw) zijn zelf slachtoffer van naammisbruik",
         "Geen motief; mogelijk onbewust van eigen EPD-vermelding; past bij geautomatiseerde naam-injectie",
         "Mousawi had wél toegang via SUBSTANCEHXQNR; Nota's employer-ID matcht",
         "Vektis AGB verifiëren; direct contact met de vier personen (via advocaat)",
         "HOOG", "Namen gebruikt als cover door een of meer systeem-actoren met EPD-schrijfrechten"),
        ("De nachtelijke batchoperaties zijn geautomatiseerde scripts, niet handmatig",
         "Exacte gelijke timestamp 03:34:55 in 7/12 documenten; 147 incrementen in 33 dagen uniform",
         "Geen bewijs van specifiek script; kan ook handmatige batch-upload zijn",
         "Git/SVN log override.css + Oracle transaction logs (Eis G/H ingebrekestelling)",
         "HOOG", "Scheduled job of extern trigger door systeembeheerder na AVG-verzoek"),
        ("Hoppinger B.V. heeft (on)bewust backdoor in productie geplaatst",
         "Hardcoded testdomein in productie-config (NB-114)",
         "Kan technische slordigheid zijn (staging URL in config vergeten)",
         "Source code audit Hoppinger contract + verantwoordelijkheidsmatrix SG",
         "MIDDEL", "Contractuele aansprakelijkheid Spaarne voor subverwerker (AVG art. 28)"),
        ("LSP/VWI bulk-pulls door onbevoegde partijen buiten behandelrelatie",
         "228 data-requests Brijder/Indigo zonder behandelrelatie (NB-113); SNOMED-scanning actief",
         "Sommige requests kunnen legitiem zijn (ketenzorg)",
         "VWI individuele AGB-log opvragen via VZVZ; Pharmeon zab-registry",
         "HOOG", "Stelselmatige surveillance buiten behandelrelatie = art. 138ab Sr"),
        ("MRK is de centrale technische actor die zowel data-injectie (2017) als data-hiding (2024) faciliteerde",
         "Zelfde persoon in toegangslog 14-12-2017 én override.css developer 2024-2025 (CC-J)",
         "Facilitair admin heeft breed systeem-toegang; correlatie ≠ causaliteit",
         "Git-log override.css (Eis H) + alle toegangslogs MRK 2017-2026",
         "HOOG", "Langdurige betrokkenheid van binnenuit — werkgeversaansprakelijkheid SG ex art. 6:170 BW"),
        ("J.M. Blauw heeft de laborders nooit zelf gezien of tekende blanco mee",
         "Geen consult in bezoekenoverzicht; laborders besteld op naam zonder klinisch contact",
         "Blauw kan telefonisch have geconsulteerd (niet gedocumenteerd)",
         "Epic-providernummer achter Blauw-laborders opvragen (toegangslog)",
         "MIDDEL", "Naam als formele autorisatie voor labs die door iemand anders werden aangevraagd"),
    ]

    for r, row in enumerate(HYPOTHESES, start=2):
        bg = LIGHT if r % 2 == 0 else WHITE
        prio = row[4]
        prio_bg = RED if prio == "HOOG" else (ORANGE if prio == "MIDDEL" else GREY)
        for c, v in enumerate(row, start=1):
            cell_bg = prio_bg if c == 5 else bg
            val(ws, r, c, v, bg=cell_bg)

    set_col_widths(ws, [50, 50, 40, 45, 8, 55])

def sheet_todo(wb, index):
    ws = wb.create_sheet("✅ TODO")
    headers = ["Prioriteit", "Deadline", "Actie", "Categorie", "NB-ref", "Status"]
    for c, h in enumerate(headers, 1):
        hdr(ws, 1, c, h)
    freeze(ws)

    TODOS = [
        ("KRITIEK", "28-06-2026", "Parnassia M26015970: portaaltoegang + logging opvragen voor vervaldatum", "Deadline", "§9-C", "OPEN"),
        ("HOOG",    "22-07-2026", "Zitting Rb Noord-Holland C/15/376914 — stukken gereed", "Rechtbank", "CC-A", "OPEN"),
        ("HOOG",    "03-08-2026", "Hof van Discipline 260153 — reageren op uitspraak", "Tuchtrecht", "BIG 47", "OPEN"),
        ("HOOG",    "PM",         "AGB J.M. Blauw via Vektis → registratiedatum vs. 29-11-2019", "Verificatie", "NB-173", "OPEN"),
        ("HOOG",    "PM",         "IGJ_VOLLEDIG_DOSSIER_20260621.docx lezen en NB's extraheren", "Analyse", "IGJ", "OPEN"),
        ("HOOG",    "PM",         "AP_VOLLEDIG_DOSSIER_20260621.docx lezen en NB's extraheren", "Analyse", "AP", "OPEN"),
        ("HOOG",    "PM",         "AP_klacht_aanvulling_driepatroon bijwerken → vierpatroon Blauw", "Brief", "NB-173", "OPEN"),
        ("HOOG",    "PM",         "Toegangslog opvragen: welk Epic-providernummer achter Blauw-laborders?", "Bewijs", "NB-173", "OPEN"),
        ("MIDDEL",  "PM",         "WERKDOCUMENT_Grothe NB-173 formeel invoegen", "Dossier", "NB-173", "OPEN"),
        ("MIDDEL",  "PM",         "Actor 470154242 identificeren (Epic-ID)", "Identificatie", "NB-170", "OPEN"),
        ("MIDDEL",  "PM",         "Caren person-ID 492769 screenshot analyseren", "Analyse", "NB-165", "OPEN"),
        ("MIDDEL",  "PM",         "NB-160 causale keten 13-01-2020 → 02-10-2024 formaliseren", "Causaal", "NB-160", "OPEN"),
        ("MIDDEL",  "PM",         "NB-148 Van der List formaliseren (bewijs Vektis + EPD label)", "Bewijs", "NB-148", "OPEN"),
        ("MIDDEL",  "PM",         "VWI individuele AGB-log via VZVZ opvragen (Mousawi/Nota)", "AVG", "NB-113", "OPEN"),
        ("MIDDEL",  "PM",         "Oracle/Caché transaction logs vorderen (Eis G ingebrekestelling)", "Juridisch", "NB-169", "OPEN"),
        ("LAAG",    "PM",         "Git/SVN log override.css incl. commits MRK/JKE (Eis H)", "Technisch", "NB-168", "OPEN"),
        ("LAAG",    "PM",         "DOC0013.XML analyseren (nog niet gelezen)", "Analyse", "NB-172", "OPEN"),
        ("LAAG",    "PM",         "HAR-bestanden mijnspaarnegasthuis.nl (84MB) analyseren", "Analyse", "NB-172", "OPEN"),
    ]

    # Also add todos from index
    for t in index.get("todos", []):
        TODOS.append(("OPEN", "PM", t, "Dossier", "", "OPEN"))

    for r, row in enumerate(TODOS, start=2):
        bg = LIGHT if r % 2 == 0 else WHITE
        prio = row[0]
        prio_bg = RED if prio == "KRITIEK" else (ORANGE if prio == "HOOG" else (YELLOW if prio == "MIDDEL" else bg))
        for c, v in enumerate(row, start=1):
            cell_bg = prio_bg if c == 1 else bg
            val(ws, r, c, v, bg=cell_bg)

    set_col_widths(ws, [10, 12, 65, 15, 12, 10])

# ── Main ───────────────────────────────────────────────────────────────────

def generate_excel(index, today):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    sheet_dashboard(wb, index, today)
    sheet_index(wb, index)
    sheet_actoren(wb, index)
    sheet_tijdlijn(wb, index)
    sheet_causaal(wb, index)
    sheet_hypotheses(wb, index)
    sheet_todo(wb, index)

    path = REPORTS / f"FORENSISCH_DOSSIER_GROTHE_{today}.xlsx"
    wb.save(path)
    return path

if __name__ == "__main__":
    import json
    from forensic.pipeline import INDEX_FILE, TODAY
    with open(INDEX_FILE) as f:
        index = json.load(f)
    p = generate_excel(index, TODAY)
    print(f"Excel: {p}")
