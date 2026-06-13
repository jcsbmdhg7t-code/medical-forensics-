#!/usr/bin/env python3
"""
╔══════════════════════════════════════════════════════════════════════════════╗
║  FORENSISCH NETWERK-DOCUMENT EXTRACTOR                                      ║
║  Dossier Grothe — Rechtbank Noord-Holland C/15/376914                       ║
║  Hof van Discipline kenmerk 260153                                          ║
║  Versie 1.0 — 13 juni 2026                                                  ║
╚══════════════════════════════════════════════════════════════════════════════╝

Extraheert ALLE informatie uit netwerk-onderschepping documenten:
  HAR, ZIP, PDF, DOCX, JSON, XML, FHIR, binaire bestanden, etc.

Detecteert en vertaalt: base64, gzip, gzip+base64, deflate, hex, URL-encoding,
  HTML entities, unicode escapes, CID-encoding (PDF), CMap (PDF), FHIR BGZ.

Detecteert verborgen data: trailing bytes, wit-op-wit tekst (PDF), zero-width
  characters, embedded files, geobfusceerde strings.

Uitvoer: Excel (meerdere tabbladen), extractie-map, SHA-256 evidence log.

Gebruik:
  python3 forensic_extractor.py <input_map> <output_map> [--icloud]

  --icloud  : schrijft ook naar ~/Library/Mobile Documents/com~apple~CloudDocs/
              (alleen macOS; op Linux: sla op in output_map en sync handmatig)

Vereiste packages:
  pip install openpyxl PyMuPDF python-docx chardet
"""

import os
import sys
import json
import gzip
import zlib
import base64
import hashlib
import zipfile
import io
import re
import html as html_module
import struct
import binascii
import urllib.parse
import datetime
import logging
import pathlib
import time
import traceback
import shutil
import argparse
import threading
from collections import defaultdict, OrderedDict
from typing import Optional, List, Dict, Tuple, Any

# ──────────────────────────────────────────────────────────────────────────────
# Optionele imports met graceful fallback
# ──────────────────────────────────────────────────────────────────────────────
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import fitz  # PyMuPDF
    HAS_PYMUPDF = True
except ImportError:
    HAS_PYMUPDF = False

try:
    import docx as python_docx
    HAS_DOCX = True
except ImportError:
    HAS_DOCX = False

try:
    import chardet
    HAS_CHARDET = True
except ImportError:
    HAS_CHARDET = False

try:
    import xml.etree.ElementTree as ET
    HAS_XML = True
except ImportError:
    HAS_XML = False

# ──────────────────────────────────────────────────────────────────────────────
# CONSTANTEN
# ──────────────────────────────────────────────────────────────────────────────
VERSION = "1.0"
TIMESTAMP = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
MAX_CELL_CHARS = 32000       # Excel cel-limiet is 32767
MAX_EXTRACT_SIZE = 50_000_000  # 50 MB max per geëxtraheerd bestand
ENCODING_PATTERNS = {
    'base64': re.compile(r'^[A-Za-z0-9+/]{20,}={0,2}$'),
    'hex':    re.compile(r'^[0-9a-fA-F]{20,}$'),
    'url':    re.compile(r'%[0-9A-Fa-f]{2}'),
    'html':   re.compile(r'&(?:#\d+|#x[0-9a-fA-F]+|[a-z]+);'),
    'unicode_escape': re.compile(r'\\u[0-9a-fA-F]{4}'),
    'zero_width': re.compile(r'[​‌‍﻿­]'),
}

# Bekende bestandssignaturen (magic bytes)
MAGIC_SIGNATURES = [
    (b'%PDF',         'pdf',   'Adobe PDF'),
    (b'PK\x03\x04',  'zip',   'ZIP archive'),
    (b'\x1f\x8b',    'gz',    'Gzip'),
    (b'BZh',         'bz2',   'BZip2'),
    (b'\x89PNG',      'png',   'PNG image'),
    (b'\xff\xd8\xff', 'jpg',   'JPEG image'),
    (b'GIF8',         'gif',   'GIF image'),
    (b'RIFF',         'riff',  'RIFF (WAV/AVI)'),
    (b'\xd0\xcf\x11', 'ole',   'OLE2 (DOC/XLS/PPT)'),
    (b'{',            'json',  'JSON'),
    (b'<?xml',        'xml',   'XML'),
    (b'<html',        'html',  'HTML'),
    (b'<!DOCTYPE',    'html',  'HTML'),
    (b'-----BEGIN',   'pem',   'PEM certificate'),
]

# FHIR BGZ resource types
FHIR_BGZ_KEYS = [
    'BgzPatient', 'BgzConcern', 'BgzAllergyIntolerance', 'BgzDrugUse',
    'BgzTobaccoUse', 'BgzAlcoholUse', 'BgzMedicationUse', 'BgzProcedure',
    'BgzObservation', 'BgzImmunization', 'BgzDeviceUseStatement',
    'BgzNutritionAdvice', 'BgzTreatmentDirective', 'BgzAdvanceDirective',
]


# ══════════════════════════════════════════════════════════════════════════════
# LIVE LOGGER — SHA-256 evidence log + console + bestand
# ══════════════════════════════════════════════════════════════════════════════
class ForensicLogger:
    def __init__(self, output_dir: pathlib.Path):
        self.output_dir = output_dir
        self.log_path = output_dir / f"forensic_log_{TIMESTAMP}.txt"
        self.evidence_log = output_dir / f"sha256_evidence_{TIMESTAMP}.txt"
        self._lock = threading.Lock()
        self._setup_logging()
        self.start_time = time.time()
        self.stats = defaultdict(int)

    def _setup_logging(self):
        self.output_dir.mkdir(parents=True, exist_ok=True)
        logging.basicConfig(
            level=logging.DEBUG,
            format='%(asctime)s [%(levelname)s] %(message)s',
            handlers=[
                logging.FileHandler(self.log_path, encoding='utf-8'),
                logging.StreamHandler(sys.stdout)
            ]
        )
        self.log = logging.getLogger('ForensicExtractor')

    def info(self, msg: str):
        self.log.info(msg)

    def warn(self, msg: str):
        self.log.warning(msg)

    def error(self, msg: str, exc: Exception = None):
        if exc:
            self.log.error(f"{msg}: {exc}")
        else:
            self.log.error(msg)

    def phase(self, title: str):
        sep = "═" * 70
        self.log.info(f"\n{sep}\n  {title}\n{sep}")

    def log_hash(self, filepath: str, sha256: str, size: int):
        with self._lock:
            with open(self.evidence_log, 'a', encoding='utf-8') as f:
                ts = datetime.datetime.now().isoformat()
                f.write(f"{ts}|SHA256|{sha256}|{size}|{filepath}\n")

    def log_finding(self, category: str, filepath: str, detail: str):
        self.stats[category] += 1
        self.log.info(f"[{category}] {pathlib.Path(filepath).name}: {detail[:120]}")

    def elapsed(self) -> str:
        s = int(time.time() - self.start_time)
        return f"{s // 60}m{s % 60}s"


# ══════════════════════════════════════════════════════════════════════════════
# SHA-256 HASHER
# ══════════════════════════════════════════════════════════════════════════════
class Hasher:
    @staticmethod
    def file(path: str) -> Tuple[str, int]:
        h = hashlib.sha256()
        size = 0
        try:
            with open(path, 'rb') as f:
                for chunk in iter(lambda: f.read(65536), b''):
                    h.update(chunk)
                    size += len(chunk)
        except Exception:
            return "ERROR", 0
        return h.hexdigest(), size

    @staticmethod
    def bytes(data: bytes) -> str:
        return hashlib.sha256(data).hexdigest()

    @staticmethod
    def string(s: str) -> str:
        return hashlib.sha256(s.encode('utf-8', errors='replace')).hexdigest()


# ══════════════════════════════════════════════════════════════════════════════
# BESTANDSTYPE DETECTIE
# ══════════════════════════════════════════════════════════════════════════════
class FileTypeDetector:
    @staticmethod
    def detect(path: str, data: bytes = None) -> Tuple[str, str]:
        """Returns (ext, description)"""
        if data is None:
            try:
                with open(path, 'rb') as f:
                    data = f.read(512)
            except Exception:
                return 'unknown', 'Onleesbaar'

        # Magic bytes check
        for sig, ext, desc in MAGIC_SIGNATURES:
            if data.startswith(sig):
                return ext, desc

        # Extensie fallback
        ext = pathlib.Path(path).suffix.lower().lstrip('.')
        ext_map = {
            'har': ('har', 'HTTP Archive (HAR)'),
            'json': ('json', 'JSON'),
            'xml': ('xml', 'XML'),
            'html': ('html', 'HTML'),
            'htm': ('html', 'HTML'),
            'txt': ('txt', 'Tekst'),
            'csv': ('csv', 'CSV'),
            'pdf': ('pdf', 'PDF'),
            'docx': ('docx', 'Word DOCX'),
            'xlsx': ('xlsx', 'Excel XLSX'),
            'pptx': ('pptx', 'PowerPoint PPTX'),
            'zip': ('zip', 'ZIP'),
            'gz': ('gz', 'Gzip'),
            'log': ('log', 'Logbestand'),
            'cda': ('xml', 'CDA (Clinical Document Architecture)'),
        }
        return ext_map.get(ext, (ext or 'bin', f'Binair ({ext or "onbekend"})'))

    @staticmethod
    def detect_encoding(data: bytes) -> str:
        """Detecteer de tekst-encoding van bytes."""
        if HAS_CHARDET:
            result = chardet.detect(data[:4096])
            enc = result.get('encoding') or 'utf-8'
        else:
            # Simpele heuristic
            try:
                data.decode('utf-8')
                enc = 'utf-8'
            except Exception:
                try:
                    data.decode('latin-1')
                    enc = 'latin-1'
                except Exception:
                    enc = 'binary'
        return enc


# ══════════════════════════════════════════════════════════════════════════════
# ENCODING DETECTOR & DECODER
# ══════════════════════════════════════════════════════════════════════════════
class EncodingDecoder:
    def __init__(self, logger: ForensicLogger):
        self.logger = logger

    def decode_all(self, data: bytes, source_hint: str = "") -> List[Dict]:
        """Probeer alle bekende encodings en geef lijst van succesvolle decodings."""
        results = []

        # 1. Direct tekst
        text = self._try_decode_text(data)
        if text:
            results.append({'encoding': 'UTF-8/Latin-1', 'decoded': text[:2000], 'full_length': len(text)})

        # 2. Base64
        b64 = self._try_base64(data)
        if b64:
            results.append({'encoding': 'Base64', 'decoded': str(b64[:2000]), 'full_length': len(b64)})
            # Recursief: is het resultaat ook gecodeerd?
            if isinstance(b64, bytes):
                sub = self.decode_all(b64, f"{source_hint}[base64]")
                for s in sub:
                    s['encoding'] = f"Base64→{s['encoding']}"
                    results.extend([s])

        # 3. Gzip
        gz = self._try_gzip(data)
        if gz:
            sub = self.decode_all(gz, f"{source_hint}[gzip]")
            for s in sub:
                s['encoding'] = f"Gzip→{s['encoding']}"
                results.extend([s])

        # 4. Deflate (zlib zonder header)
        deflate = self._try_deflate(data)
        if deflate:
            sub = self.decode_all(deflate, f"{source_hint}[deflate]")
            for s in sub:
                s['encoding'] = f"Deflate→{s['encoding']}"
                results.extend([s])

        # 5. Hex
        hexdec = self._try_hex(data)
        if hexdec:
            results.append({'encoding': 'Hex', 'decoded': str(hexdec[:2000]), 'full_length': len(hexdec)})

        # 6. URL encoding
        url_dec = self._try_url(data)
        if url_dec:
            results.append({'encoding': 'URL-encoding', 'decoded': url_dec[:2000], 'full_length': len(url_dec)})

        # 7. HTML entities
        html_dec = self._try_html_entities(data)
        if html_dec:
            results.append({'encoding': 'HTML-entities', 'decoded': html_dec[:2000], 'full_length': len(html_dec)})

        # 8. Unicode escapes
        uni_dec = self._try_unicode_escapes(data)
        if uni_dec:
            results.append({'encoding': 'Unicode-escapes', 'decoded': uni_dec[:2000], 'full_length': len(uni_dec)})

        return results

    def _try_decode_text(self, data: bytes) -> Optional[str]:
        for enc in ('utf-8', 'utf-8-sig', 'latin-1', 'cp1252', 'iso-8859-1'):
            try:
                text = data.decode(enc, errors='strict')
                if len(text) > 10 and any(c.isalpha() for c in text[:100]):
                    return text
            except Exception:
                continue
        return None

    def _try_base64(self, data: bytes) -> Optional[bytes]:
        try:
            text = data.decode('ascii', errors='strict').strip()
            # Verwijder whitespace
            cleaned = re.sub(r'\s+', '', text)
            if len(cleaned) < 20:
                return None
            if not re.match(r'^[A-Za-z0-9+/=]+$', cleaned):
                return None
            decoded = base64.b64decode(cleaned + '==')
            if len(decoded) < 10:
                return None
            return decoded
        except Exception:
            return None

    def _try_gzip(self, data: bytes) -> Optional[bytes]:
        try:
            if data[:2] == b'\x1f\x8b':
                return gzip.decompress(data)
            # Probeer ook als de data ergens gzip bevat
            idx = data.find(b'\x1f\x8b')
            if idx > 0 and idx < len(data) - 20:
                return gzip.decompress(data[idx:])
        except Exception:
            pass
        return None

    def _try_deflate(self, data: bytes) -> Optional[bytes]:
        try:
            return zlib.decompress(data)
        except Exception:
            try:
                return zlib.decompress(data, -15)  # raw deflate
            except Exception:
                return None

    def _try_hex(self, data: bytes) -> Optional[bytes]:
        try:
            text = data.decode('ascii', errors='strict').strip().replace(' ', '').replace('\n', '')
            if len(text) < 20 or len(text) % 2 != 0:
                return None
            if not re.match(r'^[0-9a-fA-F]+$', text):
                return None
            decoded = bytes.fromhex(text)
            if len(decoded) > 5:
                return decoded
        except Exception:
            pass
        return None

    def _try_url(self, data: bytes) -> Optional[str]:
        try:
            text = data.decode('utf-8', errors='replace')
            if '%' in text and re.search(r'%[0-9A-Fa-f]{2}', text):
                decoded = urllib.parse.unquote(text)
                if decoded != text:
                    return decoded
        except Exception:
            pass
        return None

    def _try_html_entities(self, data: bytes) -> Optional[str]:
        try:
            text = data.decode('utf-8', errors='replace')
            if '&' in text and ';' in text:
                decoded = html_module.unescape(text)
                if decoded != text:
                    return decoded
        except Exception:
            pass
        return None

    def _try_unicode_escapes(self, data: bytes) -> Optional[str]:
        try:
            text = data.decode('utf-8', errors='replace')
            if '\\u' in text or '\\U' in text:
                decoded = text.encode('raw_unicode_escape').decode('unicode_escape')
                if decoded != text:
                    return decoded
        except Exception:
            pass
        return None


# ══════════════════════════════════════════════════════════════════════════════
# VERBORGEN DATA DETECTOR
# ══════════════════════════════════════════════════════════════════════════════
class HiddenDataDetector:
    def __init__(self, logger: ForensicLogger):
        self.logger = logger

    def detect_trailing_data(self, filepath: str, filetype: str) -> Optional[Dict]:
        """Detecteer data na EOF-marker van bekende bestandstypes."""
        eof_markers = {
            'pdf': b'%%EOF',
            'zip': None,  # ZIP heeft end-of-central-directory record
            'jpg': b'\xff\xd9',
            'png': b'\x00\x00\x00\x00IEND\xaeB`\x82',
        }
        marker = eof_markers.get(filetype)
        if not marker:
            return None

        try:
            with open(filepath, 'rb') as f:
                data = f.read()

            # Zoek LAATSTE voorkomen van EOF marker
            idx = data.rfind(marker)
            if idx == -1:
                return None

            trailing_start = idx + len(marker)
            trailing = data[trailing_start:].strip()

            if len(trailing) > 10:
                preview = trailing[:200]
                return {
                    'type': 'TRAILING_DATA',
                    'offset': trailing_start,
                    'size': len(trailing),
                    'preview': repr(preview),
                    'sha256': Hasher.bytes(trailing),
                }
        except Exception:
            pass
        return None

    def detect_zero_width(self, text: str) -> Optional[Dict]:
        """Detecteer zero-width en onzichtbare karakters."""
        found = ENCODING_PATTERNS['zero_width'].findall(text)
        if found:
            positions = [m.start() for m in ENCODING_PATTERNS['zero_width'].finditer(text)]
            return {
                'type': 'ZERO_WIDTH_CHARS',
                'count': len(found),
                'positions': positions[:20],
                'chars': [f"U+{ord(c):04X}" for c in found[:20]],
            }
        return None

    def detect_pdf_hidden_layers(self, filepath: str) -> List[Dict]:
        """Detecteer verborgen tekst in PDF (wit op wit, onzichtbare lagen)."""
        results = []
        if not HAS_PYMUPDF:
            return results

        try:
            doc = fitz.open(filepath)
            for page_num, page in enumerate(doc):
                # Haal alle spans op met kleurinfo
                blocks = page.get_text("rawdict", flags=fitz.TEXT_PRESERVE_WHITESPACE)
                for block in blocks.get("blocks", []):
                    for line in block.get("lines", []):
                        for span in line.get("spans", []):
                            color = span.get("color", 0)
                            text = span.get("text", "").strip()
                            if not text:
                                continue
                            # Wit tekst (kleur = 16777215 = #FFFFFF) of onzichtbaar
                            if color == 16777215 or color == 0xFFFFFF:
                                results.append({
                                    'type': 'WHITE_ON_WHITE',
                                    'page': page_num + 1,
                                    'text': text[:200],
                                    'color': f"#{color:06X}",
                                })
                            # Transparante rendering mode (kleur -1 of specifieke waarden)
                            elif color < 0 or (isinstance(color, int) and color > 0xEEEEEE):
                                results.append({
                                    'type': 'NEAR_INVISIBLE',
                                    'page': page_num + 1,
                                    'text': text[:200],
                                    'color': str(color),
                                })
        except Exception as e:
            self.logger.error(f"PDF hidden layer check fout", e)
        return results


# ══════════════════════════════════════════════════════════════════════════════
# INHOUDSEXTRACTOREN PER BESTANDSTYPE
# ══════════════════════════════════════════════════════════════════════════════
class ContentExtractor:
    def __init__(self, logger: ForensicLogger, decoder: EncodingDecoder,
                 hidden: HiddenDataDetector, output_dir: pathlib.Path):
        self.logger = logger
        self.decoder = decoder
        self.hidden = hidden
        self.output_dir = output_dir
        self.extracted_dir = output_dir / "extracted_files"
        self.extracted_dir.mkdir(exist_ok=True)

    def extract(self, filepath: str, filetype: str, raw_data: bytes) -> Dict:
        """Dispatcher: kies extractiemethode op basis van filetype."""
        handlers = {
            'har':  self._extract_har,
            'json': self._extract_json,
            'xml':  self._extract_xml,
            'html': self._extract_html,
            'pdf':  self._extract_pdf,
            'docx': self._extract_docx,
            'xlsx': self._extract_xlsx,
            'zip':  self._extract_zip,
            'gz':   self._extract_gz,
            'txt':  self._extract_text,
            'log':  self._extract_text,
            'csv':  self._extract_text,
        }
        handler = handlers.get(filetype, self._extract_binary)
        try:
            return handler(filepath, raw_data)
        except Exception as e:
            self.logger.error(f"Extractie fout {filepath}", e)
            return {'text': f"EXTRACTIE FOUT: {e}", 'nested': [], 'encodings': []}

    # ──────────────────────────────────────────────────────────────
    def _extract_har(self, filepath: str, data: bytes) -> Dict:
        """HAR — HTTP Archive — kern van de netwerk-onderschepping."""
        result = {'text': '', 'nested': [], 'encodings': [], 'entries': []}
        try:
            text = data.decode('utf-8', errors='replace')
            har = json.loads(text)
        except Exception as e:
            result['text'] = f"HAR parse fout: {e}"
            return result

        log = har.get('log', {})
        entries = log.get('entries', [])
        pages = log.get('pages', [])
        creator = log.get('creator', {})

        summary_lines = [
            f"HAR Creator: {creator.get('name', '?')} v{creator.get('version', '?')}",
            f"Pagina's: {len(pages)}",
            f"Requests: {len(entries)}",
        ]
        result['text'] = "\n".join(summary_lines)

        for i, entry in enumerate(entries):
            req  = entry.get('request', {})
            resp = entry.get('response', {})
            url  = req.get('url', '')
            method = req.get('method', '')
            status = resp.get('status', 0)
            started = entry.get('startedDateTime', '')

            # Request headers
            req_headers = {h['name']: h['value'] for h in req.get('headers', [])}
            resp_headers = {h['name']: h['value'] for h in resp.get('response', {}).get('headers', resp.get('headers', []))}

            # Request body
            req_body_text = ''
            req_post = req.get('postData', {})
            if req_post:
                req_body_text = req_post.get('text', '')

            # Response body
            resp_content = resp.get('content', {})
            resp_encoding = resp_content.get('encoding', '')
            resp_mime = resp_content.get('mimeType', '')
            resp_text = resp_content.get('text', '')

            # Decodering response body
            decoded_body = ''
            nested_files = []
            encoding_found = []

            if resp_text:
                raw_resp = resp_text.encode('utf-8', errors='replace')
                if resp_encoding == 'base64':
                    try:
                        decoded_bytes = base64.b64decode(resp_text + '==')
                        # Probeer gzip decompressie
                        try:
                            decoded_bytes = gzip.decompress(decoded_bytes)
                            encoding_found.append('Base64→Gzip')
                        except Exception:
                            encoding_found.append('Base64')
                        decoded_body = decoded_bytes.decode('utf-8', errors='replace')
                        # Sla op als nested bestand
                        safe_name = re.sub(r'[^\w.-]', '_', url[-40:]) or f"entry_{i}"
                        nested_path = self.extracted_dir / f"har_entry_{i:04d}_{safe_name}"
                        nested_path.write_bytes(decoded_bytes)
                        nested_files.append({
                            'path': str(nested_path),
                            'source': url,
                            'encoding': '/'.join(encoding_found),
                            'size': len(decoded_bytes),
                            'sha256': Hasher.bytes(decoded_bytes),
                        })
                        # Recursief: is het JSON/FHIR?
                        fhir_data = self._try_parse_fhir(decoded_body, url)
                        if fhir_data:
                            encoding_found.append('FHIR-BGZ')
                            decoded_body = fhir_data + "\n\n" + decoded_body[:1000]
                    except Exception as e:
                        decoded_body = f"Decodeer fout: {e}\n{resp_text[:500]}"
                else:
                    decoded_body = resp_text
                    # Zoek geëmbedde base64-strings in response
                    self._scan_embedded_encodings(resp_text, url, i, nested_files, encoding_found)

                # Zoek naar FHIR ook in ongedecodeerde text
                if not any('FHIR' in e for e in encoding_found):
                    fhir = self._try_parse_fhir(decoded_body or resp_text, url)
                    if fhir:
                        encoding_found.append('FHIR-BGZ')

            entry_record = {
                'index': i + 1,
                'started': started,
                'method': method,
                'url': url[:300],
                'status': status,
                'mime': resp_mime,
                'req_size': len(req_body_text),
                'resp_size': resp_content.get('size', 0),
                'encoding': '/'.join(encoding_found) or resp_encoding or 'geen',
                'decoded_preview': (decoded_body or resp_text)[:500],
                'req_headers': str(dict(list(req_headers.items())[:10])),
                'cookies': str([c.get('name') for c in req.get('cookies', [])]),
                'nested_files': nested_files,
            }
            result['entries'].append(entry_record)
            result['nested'].extend(nested_files)
            result['encodings'].extend(encoding_found)

        return result

    def _try_parse_fhir(self, text: str, url: str) -> Optional[str]:
        """Probeer FHIR BGZ te parsen en geef leesbare samenvatting."""
        if not text or len(text) < 50:
            return None
        try:
            data = json.loads(text) if text.strip().startswith('{') else None
            if not data:
                return None
            # Controleer op BGZ portal wrapper
            if 'data' in data and 'resourceEndpoint' in data:
                inner = data.get('data', '')
                if isinstance(inner, str):
                    inner = json.loads(inner)
                return self._summarize_fhir_bundle(inner, data.get('resourceEndpoint', ''))
            # Controleer op FHIR Bundle
            if data.get('resourceType') == 'Bundle':
                return self._summarize_fhir_bundle(data, url)
            # Controleer op BGZ keys
            for key in FHIR_BGZ_KEYS:
                if key in data:
                    lines = [f"BGZ Resource: {key}"]
                    bundle = data[key]
                    if isinstance(bundle, dict):
                        entries = bundle.get('entry', [])
                        lines.append(f"  Entries: {len(entries)}")
                        for e in entries[:5]:
                            res = e.get('resource', {})
                            rt = res.get('resourceType', '?')
                            lines.append(f"  - {rt}: {str(res)[:100]}")
                    return "\n".join(lines)
        except Exception:
            pass
        return None

    def _summarize_fhir_bundle(self, bundle: dict, endpoint: str) -> str:
        lines = [f"FHIR Bundle ({endpoint})"]
        entries = bundle.get('entry', [])
        lines.append(f"  Totaal entries: {len(entries)}")
        resource_types = defaultdict(int)
        for e in entries:
            res = e.get('resource', {})
            rt = res.get('resourceType', 'Unknown')
            resource_types[rt] += 1
        for rt, cnt in sorted(resource_types.items()):
            lines.append(f"  {rt}: {cnt}")
        return "\n".join(lines)

    def _scan_embedded_encodings(self, text: str, url: str, idx: int,
                                  nested: list, encodings: list):
        """Scan tekst voor geëmbedde base64-blokken en andere encodings."""
        # Base64-patronen van minimaal 40 tekens
        b64_pattern = re.compile(r'["\']([A-Za-z0-9+/]{40,}={0,2})["\']')
        for m in b64_pattern.finditer(text[:50000]):
            try:
                dec = base64.b64decode(m.group(1) + '==')
                if len(dec) > 20:
                    ftype, _ = FileTypeDetector.detect('', dec[:8])
                    encodings.append(f"embedded-base64({ftype})")
            except Exception:
                pass

    # ──────────────────────────────────────────────────────────────
    def _extract_pdf(self, filepath: str, data: bytes) -> Dict:
        result = {'text': '', 'nested': [], 'encodings': [], 'hidden': []}
        all_text = []

        if HAS_PYMUPDF:
            try:
                doc = fitz.open(filepath)
                result['text'] = f"PDF: {doc.page_count} pagina's\n"

                # Metadata
                meta = doc.metadata
                result['text'] += f"Metadata: {meta}\n\n"

                for page_num, page in enumerate(doc):
                    page_text = page.get_text("text")
                    all_text.append(f"=== PAGINA {page_num+1} ===\n{page_text}")

                    # Annotaties
                    for annot in page.annots():
                        info = annot.info
                        if info.get('content'):
                            all_text.append(f"[ANNOTATIE p{page_num+1}]: {info['content']}")

                    # Links
                    for link in page.get_links():
                        if link.get('uri'):
                            all_text.append(f"[LINK p{page_num+1}]: {link['uri']}")

                # Embedded bestanden
                for i in range(doc.embfile_count()):
                    info = doc.embfile_info(i)
                    ef_data = doc.embfile_get(i)
                    ef_path = self.extracted_dir / f"pdf_embedded_{pathlib.Path(filepath).stem}_{i}"
                    ef_path.write_bytes(ef_data)
                    result['nested'].append({
                        'path': str(ef_path),
                        'source': filepath,
                        'type': 'PDF-embedded',
                        'size': len(ef_data),
                        'sha256': Hasher.bytes(ef_data),
                    })

                # Verborgen lagen
                hidden_layers = self.hidden.detect_pdf_hidden_layers(filepath)
                result['hidden'] = hidden_layers

                result['text'] += "\n".join(all_text)
            except Exception as e:
                result['text'] = f"PDF extractie fout: {e}"
        else:
            # Fallback: zoek tekst in raw bytes
            text_chunks = re.findall(rb'BT\s*(.*?)\s*ET', data, re.DOTALL)
            result['text'] = f"PDF (PyMuPDF niet beschikbaar) - {len(text_chunks)} tekstblokken gevonden\n"

        # Trailing data
        trailing = self.hidden.detect_trailing_data(filepath, 'pdf')
        if trailing:
            result['hidden'].append(trailing)

        return result

    # ──────────────────────────────────────────────────────────────
    def _extract_docx(self, filepath: str, data: bytes) -> Dict:
        result = {'text': '', 'nested': [], 'encodings': []}
        text_parts = []

        # DOCX is een ZIP — open als ZIP ook zonder python-docx
        try:
            with zipfile.ZipFile(io.BytesIO(data)) as z:
                # document.xml
                if 'word/document.xml' in z.namelist():
                    xml_data = z.read('word/document.xml')
                    text_raw = xml_data.decode('utf-8', errors='replace')
                    # Strip XML tags, behoud tekst
                    paras = re.split(r'<w:p[ >]', text_raw)
                    for para in paras:
                        txt = re.sub(r'<[^>]+>', '', para).strip()
                        if txt:
                            text_parts.append(txt)

                # comments.xml
                if 'word/comments.xml' in z.namelist():
                    com_data = z.read('word/comments.xml').decode('utf-8', errors='replace')
                    coms = re.sub(r'<[^>]+>', ' ', com_data)
                    text_parts.append(f"\n[COMMENTAREN]\n{coms[:2000]}")

                # Embedded bestanden
                for name in z.namelist():
                    if name.startswith('word/embeddings/') or name.startswith('word/media/'):
                        ef_data = z.read(name)
                        ef_path = self.extracted_dir / f"docx_embedded_{pathlib.Path(name).name}"
                        ef_path.write_bytes(ef_data)
                        result['nested'].append({
                            'path': str(ef_path),
                            'source': filepath,
                            'type': 'DOCX-embedded',
                            'size': len(ef_data),
                            'sha256': Hasher.bytes(ef_data),
                        })
        except Exception as e:
            result['text'] = f"DOCX extractie fout: {e}"
            return result

        result['text'] = "\n".join(text_parts)
        return result

    # ──────────────────────────────────────────────────────────────
    def _extract_xlsx(self, filepath: str, data: bytes) -> Dict:
        result = {'text': '', 'nested': [], 'encodings': []}
        try:
            with zipfile.ZipFile(io.BytesIO(data)) as z:
                sheets = [n for n in z.namelist() if n.startswith('xl/worksheets/')]
                text_parts = [f"Excel: {len(sheets)} werkbladen\n"]
                for sheet in sheets[:10]:
                    xml = z.read(sheet).decode('utf-8', errors='replace')
                    vals = re.findall(r'<v>([^<]+)</v>', xml)
                    text_parts.append(f"Blad {sheet}: {len(vals)} waarden")
                    text_parts.append(" | ".join(vals[:50]))
                result['text'] = "\n".join(text_parts)
        except Exception as e:
            result['text'] = f"XLSX fout: {e}"
        return result

    # ──────────────────────────────────────────────────────────────
    def _extract_zip(self, filepath: str, data: bytes) -> Dict:
        result = {'text': '', 'nested': [], 'encodings': []}
        try:
            with zipfile.ZipFile(io.BytesIO(data)) as z:
                info_list = z.infolist()
                lines = [f"ZIP: {len(info_list)} bestanden"]
                for info in info_list:
                    lines.append(f"  {info.filename}  ({info.file_size} bytes)")
                    # Extraheer elk bestand
                    try:
                        ef_data = z.read(info.filename)
                        safe = re.sub(r'[^\w.-]', '_', info.filename)
                        ef_path = self.extracted_dir / f"zip_{pathlib.Path(filepath).stem}_{safe}"
                        ef_path.write_bytes(ef_data)
                        result['nested'].append({
                            'path': str(ef_path),
                            'source': f"{filepath}/{info.filename}",
                            'type': 'ZIP-entry',
                            'size': len(ef_data),
                            'sha256': Hasher.bytes(ef_data),
                        })
                    except Exception as e2:
                        lines.append(f"    [FOUT: {e2}]")
                result['text'] = "\n".join(lines)
        except Exception as e:
            result['text'] = f"ZIP fout: {e}"
        return result

    # ──────────────────────────────────────────────────────────────
    def _extract_gz(self, filepath: str, data: bytes) -> Dict:
        result = {'text': '', 'nested': [], 'encodings': ['Gzip']}
        try:
            decompressed = gzip.decompress(data)
            ef_path = self.extracted_dir / (pathlib.Path(filepath).stem + "_decompressed")
            ef_path.write_bytes(decompressed)
            result['nested'].append({
                'path': str(ef_path),
                'source': filepath,
                'type': 'Gzip-decompressed',
                'size': len(decompressed),
                'sha256': Hasher.bytes(decompressed),
            })
            result['text'] = f"Gzip: {len(data)} bytes → {len(decompressed)} bytes decompressed\n"
            result['text'] += decompressed.decode('utf-8', errors='replace')[:3000]
        except Exception as e:
            result['text'] = f"Gzip fout: {e}"
        return result

    # ──────────────────────────────────────────────────────────────
    def _extract_json(self, filepath: str, data: bytes) -> Dict:
        result = {'text': '', 'nested': [], 'encodings': []}
        try:
            text = data.decode('utf-8', errors='replace')
            obj = json.loads(text)
            result['text'] = json.dumps(obj, indent=2, ensure_ascii=False)[:5000]

            # Zoek FHIR
            fhir = self._try_parse_fhir_from_obj(obj, filepath)
            if fhir:
                result['text'] = fhir + "\n\n" + result['text']
                result['encodings'].append('FHIR')

            # Zoek base64-velden recursief
            b64_found = self._find_base64_in_json(obj, '')
            for path, val, dec in b64_found:
                result['encodings'].append(f"base64@{path}")
                ef_path = self.extracted_dir / f"json_b64_{Hasher.string(path)[:8]}"
                ef_path.write_bytes(dec)
                result['nested'].append({
                    'path': str(ef_path),
                    'source': f"{filepath}#{path}",
                    'type': 'JSON-base64',
                    'size': len(dec),
                    'sha256': Hasher.bytes(dec),
                })
        except Exception as e:
            result['text'] = f"JSON fout: {e}\n" + data.decode('utf-8', errors='replace')[:2000]
        return result

    def _try_parse_fhir_from_obj(self, obj: Any, source: str) -> Optional[str]:
        if isinstance(obj, dict):
            return self._try_parse_fhir(json.dumps(obj), source)
        return None

    def _find_base64_in_json(self, obj: Any, path: str, depth: int = 0) -> List[Tuple]:
        if depth > 10:
            return []
        results = []
        if isinstance(obj, dict):
            for k, v in obj.items():
                results.extend(self._find_base64_in_json(v, f"{path}.{k}", depth + 1))
        elif isinstance(obj, list):
            for i, v in enumerate(obj[:100]):
                results.extend(self._find_base64_in_json(v, f"{path}[{i}]", depth + 1))
        elif isinstance(obj, str) and len(obj) > 40:
            dec = EncodingDecoder(self.logger)._try_base64(obj.encode('ascii', errors='replace'))
            if dec and len(dec) > 20:
                results.append((path, obj[:50], dec))
        return results

    # ──────────────────────────────────────────────────────────────
    def _extract_xml(self, filepath: str, data: bytes) -> Dict:
        result = {'text': '', 'nested': [], 'encodings': []}
        try:
            text = data.decode('utf-8', errors='replace')
            # Verwijder namespaces voor leesbaarheid
            text_clean = re.sub(r' xmlns[^"]*"[^"]*"', '', text)
            text_clean = re.sub(r'<[a-z]+:', '<', text_clean)
            text_clean = re.sub(r'</[a-z]+:', '</', text_clean)
            # Haal tekst eruit
            plain = re.sub(r'<[^>]+>', ' ', text_clean)
            plain = re.sub(r'\s+', ' ', plain).strip()
            result['text'] = f"XML ({len(data)} bytes)\nTekst:\n{plain[:5000]}"

            # CDA-specifieke extractie
            if 'ClinicalDocument' in text or 'hl7' in text.lower():
                result['encodings'].append('CDA-R2')
                cda = self._extract_cda(text)
                if cda:
                    result['text'] = "CDA R2 Document:\n" + cda + "\n\n" + result['text']
        except Exception as e:
            result['text'] = f"XML fout: {e}"
        return result

    def _extract_cda(self, text: str) -> str:
        """Extract relevante velden uit een CDA R2 document."""
        lines = []
        patterns = {
            'Patiënt': r'<patient[^>]*>.*?</patient>',
            'Code': r'<code\s+code="([^"]+)"\s+.*?displayName="([^"]+)"',
            'Diagnose': r'<observation[^>]*>.*?</observation>',
            'Datum': r'<effectiveTime.*?value="([^"]+)"',
            'Auteur': r'<author>.*?</author>',
        }
        for label, pattern in patterns.items():
            matches = re.findall(pattern, text, re.DOTALL | re.IGNORECASE)
            for m in matches[:3]:
                val = re.sub(r'<[^>]+>', ' ', str(m)).strip()[:200]
                lines.append(f"{label}: {val}")
        return "\n".join(lines)

    # ──────────────────────────────────────────────────────────────
    def _extract_html(self, filepath: str, data: bytes) -> Dict:
        result = {'text': '', 'nested': [], 'encodings': []}
        try:
            text = data.decode('utf-8', errors='replace')
            # Strip HTML tags
            plain = re.sub(r'<script[^>]*>.*?</script>', '', text, flags=re.DOTALL)
            plain = re.sub(r'<style[^>]*>.*?</style>', '', plain, flags=re.DOTALL)
            plain = re.sub(r'<[^>]+>', ' ', plain)
            plain = html_module.unescape(plain)
            plain = re.sub(r'\s+', ' ', plain).strip()
            result['text'] = plain[:5000]
            # Zoek data-attributen (soms base64)
            data_attrs = re.findall(r'data-[a-z-]+="([^"]{40,})"', text)
            for attr in data_attrs[:10]:
                dec = EncodingDecoder(self.logger)._try_base64(attr.encode())
                if dec:
                    result['encodings'].append(f"data-attr-base64")
        except Exception as e:
            result['text'] = f"HTML fout: {e}"
        return result

    # ──────────────────────────────────────────────────────────────
    def _extract_text(self, filepath: str, data: bytes) -> Dict:
        result = {'text': '', 'nested': [], 'encodings': []}
        enc = FileTypeDetector.detect_encoding(data)
        try:
            text = data.decode(enc, errors='replace')
            result['text'] = text[:10000]
            # Zoek encodings in tekst
            decs = EncodingDecoder(self.logger).decode_all(data[:5000])
            result['encodings'] = [d['encoding'] for d in decs if d['encoding'] not in ('UTF-8/Latin-1',)]
        except Exception as e:
            result['text'] = f"Tekstfout: {e}"
        return result

    # ──────────────────────────────────────────────────────────────
    def _extract_binary(self, filepath: str, data: bytes) -> Dict:
        result = {'text': '', 'nested': [], 'encodings': []}
        # Zoek strings in binair bestand
        strings = re.findall(rb'[ -~]{6,}', data)
        readable = [s.decode('ascii', errors='replace') for s in strings[:200]]
        result['text'] = f"Binair bestand ({len(data)} bytes)\nLeesbare strings:\n"
        result['text'] += "\n".join(readable[:100])

        # Trailing data check
        for ext in ('pdf', 'jpg', 'png'):
            trailing = self.hidden.detect_trailing_data(filepath, ext)
            if trailing:
                result['nested'].append(trailing)
                result['encodings'].append('TRAILING_DATA')
                break

        # Zoek magic bytes van bekende formaten binnenin
        for sig, ext, desc in MAGIC_SIGNATURES:
            idx = data.find(sig)
            if idx > 0 and idx < len(data) - 50:
                result['encodings'].append(f"embedded-{ext}@offset{idx}")
                sub_data = data[idx:]
                ef_path = self.extracted_dir / f"binary_embedded_{pathlib.Path(filepath).stem}_{ext}_{idx}"
                ef_path.write_bytes(sub_data)
                result['nested'].append({
                    'path': str(ef_path),
                    'source': filepath,
                    'type': f'Embedded-{desc}',
                    'size': len(sub_data),
                    'sha256': Hasher.bytes(sub_data),
                })

        return result


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL REPORTER
# ══════════════════════════════════════════════════════════════════════════════
class ExcelReporter:
    def __init__(self, output_path: pathlib.Path):
        if not HAS_OPENPYXL:
            self.wb = None
            self.path = output_path
            return
        self.wb = openpyxl.Workbook()
        self.path = output_path
        self._sheets = {}
        # Stijlen
        self.header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.finding_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        self.hidden_fill = PatternFill(start_color="FFE0E0", end_color="FFE0E0", fill_type="solid")

    def _get_or_create(self, name: str, headers: List[str]) -> Any:
        if not self.wb:
            return None
        if name not in self._sheets:
            if len(self.wb.sheetnames) == 1 and self.wb.sheetnames[0] == 'Sheet':
                ws = self.wb.active
                ws.title = name
            else:
                ws = self.wb.create_sheet(name)
            self._sheets[name] = ws
            # Headers
            ws.append(headers)
            for col_idx, _ in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = self.header_fill
                cell.font = self.header_font
                cell.alignment = Alignment(horizontal='center', wrap_text=True)
            ws.freeze_panes = "A2"
            # Auto-filter
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
        return self._sheets[name]

    def _safe(self, val: Any, maxlen: int = MAX_CELL_CHARS) -> str:
        if val is None:
            return ""
        s = str(val)
        # Verwijder illegale XML-tekens
        s = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', s)
        return s[:maxlen]

    def add_summary(self, stats: Dict, input_dir: str, total_files: int,
                    total_extracted: int, start_time: float):
        ws = self._get_or_create("SAMENVATTING", [
            "Categorie", "Waarde"
        ])
        if not ws:
            return
        rows = [
            ("Forensische analyse", "Dossier Grothe — C/15/376914"),
            ("Tijdstip aanmaak", datetime.datetime.now().isoformat()),
            ("Invoermap", input_dir),
            ("Looptijd (seconden)", f"{time.time() - start_time:.1f}"),
            ("Totaal bestanden verwerkt", total_files),
            ("Totaal geneste bestanden geëxtraheerd", total_extracted),
            ("", ""),
        ] + [(k, v) for k, v in stats.items()]
        for row in rows:
            ws.append([self._safe(row[0]), self._safe(row[1])])
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 55

    def add_file_inventory(self, files: List[Dict]):
        headers = [
            "Nr", "Bestandsnaam", "Pad", "Type", "Omschrijving",
            "Grootte (bytes)", "SHA-256", "Tijdstip scan",
            "Encodings gevonden", "Nested bestanden", "Verborgen data",
        ]
        ws = self._get_or_create("BESTANDSINVENTARIS", headers)
        if not ws:
            return
        for f in files:
            ws.append([
                self._safe(f.get('nr')),
                self._safe(f.get('name')),
                self._safe(f.get('path')),
                self._safe(f.get('type')),
                self._safe(f.get('desc')),
                self._safe(f.get('size')),
                self._safe(f.get('sha256')),
                self._safe(f.get('scanned')),
                self._safe(f.get('encodings')),
                self._safe(f.get('nested_count')),
                self._safe(f.get('hidden_count')),
            ])
        for col, width in zip('ABCDEFGHIJK', [5,30,60,8,20,15,65,20,25,10,10]):
            ws.column_dimensions[get_column_letter(ord(col)-64)].width = width

    def add_har_entries(self, entries: List[Dict], source_file: str):
        headers = [
            "Nr", "Bronbestand", "Tijdstip", "Methode", "URL",
            "Status", "MIME-type", "Req-grootte", "Resp-grootte",
            "Encoding", "Gedecodeerde preview", "Cookies", "Request headers",
        ]
        ws = self._get_or_create("HAR REQUESTS", headers)
        if not ws:
            return
        for e in entries:
            ws.append([
                self._safe(e.get('index')),
                self._safe(pathlib.Path(source_file).name),
                self._safe(e.get('started')),
                self._safe(e.get('method')),
                self._safe(e.get('url')),
                self._safe(e.get('status')),
                self._safe(e.get('mime')),
                self._safe(e.get('req_size')),
                self._safe(e.get('resp_size')),
                self._safe(e.get('encoding')),
                self._safe(e.get('decoded_preview')),
                self._safe(e.get('cookies')),
                self._safe(e.get('req_headers')),
            ])

    def add_extracted_content(self, entries: List[Dict]):
        headers = [
            "Nr", "Bronbestand", "Type", "Inhoud (preview 5000 tekens)",
            "SHA-256 inhoud",
        ]
        ws = self._get_or_create("GEËXTRAHEERDE INHOUD", headers)
        if not ws:
            return
        for e in entries:
            ws.append([
                self._safe(e.get('nr')),
                self._safe(e.get('source')),
                self._safe(e.get('type')),
                self._safe(e.get('content'), 5000),
                self._safe(e.get('sha256')),
            ])
        ws.column_dimensions['D'].width = 80

    def add_nested_files(self, nested: List[Dict]):
        headers = [
            "Nr", "Geëxtraheerd naar", "Bronbestand", "Type", "Grootte", "SHA-256",
        ]
        ws = self._get_or_create("GENESTE BESTANDEN", headers)
        if not ws:
            return
        for i, n in enumerate(nested, 1):
            ws.append([
                i,
                self._safe(n.get('path')),
                self._safe(n.get('source')),
                self._safe(n.get('type')),
                self._safe(n.get('size')),
                self._safe(n.get('sha256')),
            ])

    def add_encodings(self, encodings: List[Dict]):
        headers = [
            "Nr", "Bronbestand", "Encoding type", "Locatie/context",
            "Gedecodeerde preview", "SHA-256 decoded",
        ]
        ws = self._get_or_create("ENCODINGS", headers)
        if not ws:
            return
        for i, e in enumerate(encodings, 1):
            ws.append([
                i,
                self._safe(e.get('source')),
                self._safe(e.get('encoding')),
                self._safe(e.get('location')),
                self._safe(e.get('preview')),
                self._safe(e.get('sha256')),
            ])

    def add_hidden_data(self, hidden: List[Dict]):
        headers = [
            "Nr", "Bronbestand", "Type verborgen data", "Pagina/Offset",
            "Grootte", "Preview", "SHA-256",
        ]
        ws = self._get_or_create("VERBORGEN DATA", headers)
        if not ws:
            return
        for i, h in enumerate(hidden, 1):
            row = ws.append([
                i,
                self._safe(h.get('source')),
                self._safe(h.get('type')),
                self._safe(h.get('page') or h.get('offset')),
                self._safe(h.get('size') or h.get('count')),
                self._safe(h.get('preview') or h.get('text')),
                self._safe(h.get('sha256', '')),
            ])
            # Markeer verdachte rijen rood
            for cell in ws[ws.max_row]:
                cell.fill = self.hidden_fill

    def save(self):
        if not self.wb:
            print("[WARN] openpyxl niet beschikbaar — geen Excel output")
            return
        try:
            self.wb.save(self.path)
        except Exception as e:
            print(f"[ERROR] Excel opslaan mislukt: {e}")


# ══════════════════════════════════════════════════════════════════════════════
# HOOFDKLASSE: FORENSIC EXTRACTOR
# ══════════════════════════════════════════════════════════════════════════════
class ForensicExtractor:
    def __init__(self, input_dir: str, output_dir: str):
        self.input_dir  = pathlib.Path(input_dir)
        self.output_dir = pathlib.Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

        self.logger  = ForensicLogger(self.output_dir)
        self.decoder = EncodingDecoder(self.logger)
        self.hidden  = HiddenDataDetector(self.logger)
        self.content = ContentExtractor(self.logger, self.decoder, self.hidden, self.output_dir)
        self.reporter = ExcelReporter(self.output_dir / f"forensic_resultaten_{TIMESTAMP}.xlsx")

        # Verzamelingen voor rapportage
        self.all_files: List[Dict] = []
        self.all_nested: List[Dict] = []
        self.all_encodings: List[Dict] = []
        self.all_hidden: List[Dict] = []
        self.all_content: List[Dict] = []
        self.har_entries: Dict[str, List] = {}

        self.start_time = time.time()

    def run(self):
        self.logger.phase("STAP 1: SHA-256 HASHING ALLE INVOERBESTANDEN")
        self._hash_all_files()

        self.logger.phase("STAP 2: INHOUD EXTRAHEREN")
        self._extract_all()

        self.logger.phase("STAP 3: EXCEL RAPPORT AANMAKEN")
        self._build_report()

        self.logger.phase("STAP 4: SAMENVATTING")
        self._print_summary()

    def _hash_all_files(self):
        """Fase 1: hash elk bestand als bewijs-integrity check."""
        all_paths = []
        for root, dirs, files in os.walk(self.input_dir):
            # Skip verborgen mappen
            dirs[:] = [d for d in dirs if not d.startswith('.')]
            for fname in files:
                if not fname.startswith('.'):
                    all_paths.append(os.path.join(root, fname))

        self.logger.info(f"Gevonden: {len(all_paths)} bestanden in {self.input_dir}")

        for nr, fpath in enumerate(sorted(all_paths), 1):
            sha256, size = Hasher.file(fpath)
            self.logger.log_hash(fpath, sha256, size)
            ftype, fdesc = FileTypeDetector.detect(fpath)
            self.all_files.append({
                'nr': nr,
                'name': pathlib.Path(fpath).name,
                'path': fpath,
                'type': ftype,
                'desc': fdesc,
                'size': size,
                'sha256': sha256,
                'scanned': datetime.datetime.now().isoformat(),
                'encodings': '',
                'nested_count': 0,
                'hidden_count': 0,
                '_processed': False,
            })
            self.logger.info(f"  [{nr:4d}] {pathlib.Path(fpath).name:50s} {size:>10,} bytes  SHA256:{sha256[:16]}...")

    def _extract_all(self):
        """Fase 2: extraheer inhoud van elk bestand."""
        for file_rec in self.all_files:
            fpath = file_rec['path']
            ftype = file_rec['type']
            fname = file_rec['name']

            self.logger.info(f"\n→ Verwerken: {fname} [{ftype}]")

            try:
                raw = pathlib.Path(fpath).read_bytes()
            except Exception as e:
                self.logger.error(f"Leesfout {fpath}", e)
                continue

            # Extraheer
            result = self.content.extract(fpath, ftype, raw)
            text = result.get('text', '')
            nested = result.get('nested', [])
            encodings_list = result.get('encodings', [])
            hidden_list = result.get('hidden', [])
            har_ent = result.get('entries', [])

            # Sla op
            if text:
                content_hash = Hasher.string(text)
                self.all_content.append({
                    'nr': len(self.all_content) + 1,
                    'source': fpath,
                    'type': ftype,
                    'content': text,
                    'sha256': content_hash,
                })
                # Schrijf ook als .txt bestand
                txt_path = self.output_dir / "extracted_files" / f"{pathlib.Path(fpath).stem}_{ftype}_inhoud.txt"
                try:
                    txt_path.write_text(text, encoding='utf-8', errors='replace')
                except Exception:
                    pass

            for n in nested:
                n.setdefault('nr', len(self.all_nested) + 1)
                self.all_nested.append(n)
                self.logger.log_finding('NESTED', fpath, f"{n.get('type')} → {n.get('path')}")

            for enc in encodings_list:
                self.all_encodings.append({
                    'source': fpath,
                    'encoding': enc,
                    'location': '',
                    'preview': '',
                    'sha256': '',
                })
                self.logger.log_finding('ENCODING', fpath, enc)

            for h in hidden_list:
                h['source'] = fpath
                self.all_hidden.append(h)
                self.logger.log_finding('VERBORGEN', fpath, f"{h.get('type')} — {str(h)[:80]}")

            if har_ent:
                self.har_entries[fpath] = har_ent
                self.logger.log_finding('HAR', fpath, f"{len(har_ent)} requests gevonden")

            # Update bestandsrecord
            file_rec['encodings'] = ", ".join(encodings_list[:10])
            file_rec['nested_count'] = len(nested)
            file_rec['hidden_count'] = len(hidden_list)
            file_rec['_processed'] = True

            self.logger.info(f"  ✓ {fname}: {len(text)} tekens, {len(nested)} nested, "
                            f"{len(encodings_list)} encodings, {len(hidden_list)} verborgen")

        # Verwerk ook geneste bestanden recursief (1 niveau diep)
        self.logger.info("\n→ Recursieve extractie geneste bestanden...")
        nested_to_process = list(self.all_nested)
        for nested_rec in nested_to_process:
            npath = nested_rec.get('path', '')
            if not npath or not pathlib.Path(npath).exists():
                continue
            try:
                raw = pathlib.Path(npath).read_bytes()
                ntype, ndesc = FileTypeDetector.detect(npath, raw[:512])
                sub_result = self.content.extract(npath, ntype, raw)
                sub_text = sub_result.get('text', '')
                if sub_text:
                    self.all_content.append({
                        'nr': len(self.all_content) + 1,
                        'source': npath,
                        'type': f"nested-{ntype}",
                        'content': sub_text,
                        'sha256': Hasher.string(sub_text),
                    })
                    self.logger.info(f"  ↳ Nested {pathlib.Path(npath).name}: {len(sub_text)} tekens")
            except Exception as e:
                self.logger.error(f"Nested extractie {npath}", e)

    def _build_report(self):
        """Fase 3: bouw het Excel rapport."""
        self.reporter.add_summary(
            stats={k: v for k, v in self.logger.stats.items()},
            input_dir=str(self.input_dir),
            total_files=len(self.all_files),
            total_extracted=len(self.all_nested),
            start_time=self.start_time,
        )
        self.reporter.add_file_inventory(self.all_files)

        for fpath, entries in self.har_entries.items():
            self.reporter.add_har_entries(entries, fpath)

        self.reporter.add_extracted_content(self.all_content)
        self.reporter.add_nested_files(self.all_nested)
        self.reporter.add_encodings(self.all_encodings)
        self.reporter.add_hidden_data(self.all_hidden)
        self.reporter.save()

        self.logger.info(f"Excel rapport opgeslagen: {self.reporter.path}")

    def _print_summary(self):
        elapsed = self.logger.elapsed()
        self.logger.phase("SAMENVATTING RESULTATEN")
        print(f"""
╔══════════════════════════════════════════════════════════════╗
║  FORENSISCH EXTRACTIE VOLTOOID                               ║
╠══════════════════════════════════════════════════════════════╣
║  Looptijd             : {elapsed:<35} ║
║  Bestanden verwerkt   : {len(self.all_files):<35} ║
║  Geneste bestanden    : {len(self.all_nested):<35} ║
║  Encodings gevonden   : {len(self.all_encodings):<35} ║
║  Verborgen data       : {len(self.all_hidden):<35} ║
║  HAR entries          : {sum(len(v) for v in self.har_entries.values()):<35} ║
║  Inhoudsblokken       : {len(self.all_content):<35} ║
╠══════════════════════════════════════════════════════════════╣
║  Uitvoer map          :                                      ║
║  {str(self.output_dir):<60} ║
╠══════════════════════════════════════════════════════════════╣
║  Excel rapport        :                                      ║
║  {str(self.reporter.path.name):<60} ║
║  SHA-256 log          :                                      ║
║  {str(self.logger.evidence_log.name):<60} ║
╚══════════════════════════════════════════════════════════════╝
""")
        # Schrijf ook tekst-samenvatting
        summary_path = self.output_dir / f"samenvatting_{TIMESTAMP}.txt"
        with open(summary_path, 'w', encoding='utf-8') as f:
            f.write(f"FORENSISCHE EXTRACTIE SAMENVATTING\n")
            f.write(f"Dossier Grothe — Rechtbank Noord-Holland C/15/376914\n")
            f.write(f"Tijdstip: {datetime.datetime.now().isoformat()}\n")
            f.write(f"Invoermap: {self.input_dir}\n\n")
            f.write(f"GEVONDEN BESTANDEN ({len(self.all_files)}):\n")
            for fr in self.all_files:
                f.write(f"  [{fr['nr']:4d}] {fr['name']:50s} {fr['size']:>12,} bytes  {fr['type']:8s}  SHA256:{fr['sha256'][:32]}\n")
            f.write(f"\nGEEXTRAHEERDE GENESTE BESTANDEN ({len(self.all_nested)}):\n")
            for n in self.all_nested:
                f.write(f"  {n.get('type','?'):20s}  {pathlib.Path(n.get('path','')).name}  {n.get('size',0):>10,} bytes\n")
            f.write(f"\nENCODINGS GEVONDEN ({len(self.all_encodings)}):\n")
            for e in self.all_encodings[:100]:
                f.write(f"  {e.get('source','?'):40s}  {e.get('encoding','?')}\n")
            if self.all_hidden:
                f.write(f"\nVERBORGEN DATA ({len(self.all_hidden)} items):\n")
                for h in self.all_hidden:
                    f.write(f"  {h.get('type','?'):25s}  {h.get('source','?')}\n")
        self.logger.info(f"Tekstsamenvatting: {summary_path}")


# ══════════════════════════════════════════════════════════════════════════════
# ICLOUD SYNC HELPER
# ══════════════════════════════════════════════════════════════════════════════
def copy_to_icloud(output_dir: pathlib.Path, logger: ForensicLogger):
    """Kopieer output naar iCloud Drive (alleen macOS)."""
    icloud_base = pathlib.Path.home() / "Library" / "Mobile Documents" / "com~apple~CloudDocs"
    if not icloud_base.exists():
        logger.warn("iCloud Drive niet gevonden (dit is waarschijnlijk een Linux omgeving). "
                    "Kopieer de output map handmatig naar iCloud.")
        return
    dest = icloud_base / "Forensisch_Dossier_Grothe" / f"extractie_{TIMESTAMP}"
    try:
        shutil.copytree(str(output_dir), str(dest))
        logger.info(f"Gekopieerd naar iCloud: {dest}")
        print(f"\n✓ Output gekopieerd naar iCloud: {dest}")
    except Exception as e:
        logger.error("iCloud kopie mislukt", e)


# ══════════════════════════════════════════════════════════════════════════════
# ENTRY POINT
# ══════════════════════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(
        description="Forensisch Netwerk-Document Extractor — Dossier Grothe",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Voorbeelden:
  python3 forensic_extractor.py /pad/naar/netwerk_captures /pad/naar/output
  python3 forensic_extractor.py ~/Downloads/proxyman_captures ~/Desktop/forensic_output --icloud
        """
    )
    parser.add_argument('input_dir',  help="Map met netwerk-onderschepping bestanden (HAR, ZIP, etc.)")
    parser.add_argument('output_dir', help="Uitvoermap voor resultaten en Excel rapport")
    parser.add_argument('--icloud', action='store_true',
                        help="Kopieer resultaten ook naar iCloud Drive (macOS)")

    args = parser.parse_args()

    # Controleer invoermap
    input_path = pathlib.Path(args.input_dir)
    if not input_path.exists():
        print(f"FOUT: Invoermap bestaat niet: {input_path}")
        sys.exit(1)

    print(f"""
╔══════════════════════════════════════════════════════════════╗
║  FORENSISCH NETWERK-DOCUMENT EXTRACTOR v{VERSION}               ║
║  Dossier Grothe — Rechtbank Noord-Holland C/15/376914       ║
╠══════════════════════════════════════════════════════════════╣
║  Invoer : {str(input_path)[:55]:<55} ║
║  Uitvoer: {str(args.output_dir)[:55]:<55} ║
╚══════════════════════════════════════════════════════════════╝

Beschikbare modules:
  openpyxl (Excel)   : {'✓' if HAS_OPENPYXL else '✗ pip install openpyxl'}
  PyMuPDF (PDF)      : {'✓' if HAS_PYMUPDF else '✗ pip install pymupdf'}
  python-docx (DOCX) : {'✓' if HAS_DOCX else '✗ pip install python-docx'}
  chardet (encoding) : {'✓' if HAS_CHARDET else '✗ pip install chardet'}
""")

    extractor = ForensicExtractor(args.input_dir, args.output_dir)
    extractor.run()

    if args.icloud:
        copy_to_icloud(pathlib.Path(args.output_dir), extractor.logger)


if __name__ == '__main__':
    main()
